import openpyxl
from openpyxl import Workbook
import pathlib

arquivo = pathlib.Path("src/db/teste.xlsx")
if arquivo.exists():
    workbook = openpyxl.load_workbook("src/db/teste.xlsx")
    if "Sistemas" in workbook.sheetnames:
        sistemas = workbook["Sistemas"]
    else:
        sistemas = workbook.create_sheet("Sistemas")
        sistemas['A1'] = "Código do sistema"
        sistemas['B1'] = "Nome do sistema"
        
else:
    workbook = Workbook()
    sistemas = workbook.active
    sistemas.title = "Sistemas"
    sistemas['A1'] = "Código do sistema"
    sistemas['B1'] = "Nome do sistema"     
    workbook.save("src/db/teste.xlsx")


def criar_sistema():
    for row in sistemas.iter_rows(min_row=2, max_row=5, max_col=1, values_only= True):
        for celula in row:
            print(celula.internal_value)
    idsistema = input("Código do sistema: ")
    while len(idsistema) > 15:
        idsistema = input("Código do sistema deve ser menor que 15 caracteres: ")
    print("Id inserido com sucesso")

    nomesistema = input("Nome do Sistema: ")
    while len(nomesistema) > 20:
        nomesistema = input("O nome do sistema deve ser menor que 20 caracteres: ")
    print("Id inserido com sucesso")

    try:
        sistemas.cell(column=1, row=sistemas.max_row+1, value=idsistema)
        sistemas.cell(column=2, row=sistemas.max_row, value=nomesistema)

        workbook.save(r"src/db/teste.xlsx")
        print("Dados salvos")
    except Exception as e:
        print("Ocorreu um erro", e)

criar_sistema()